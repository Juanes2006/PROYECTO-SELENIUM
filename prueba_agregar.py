from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook
import time



# ---------- INICIO REPORTE .txt ----------

reporte = open("reporte_agregar_usuarios.txt", "w", encoding="utf-8")
reporte.write("=== REPORTE AUTOMATIZACIÓN SELENIUM ===\n\n")




# ---------- 1. ABRIR NAVEGADOR ----------
driver = webdriver.Chrome()
driver.get("http://localhost:8000/iniciarSesion")

login_username = driver.find_element(By.NAME, "username")
login_password = driver.find_element(By.NAME, "password")


login_username.send_keys("je2006")
login_password.send_keys("Manila2026")
login_password.send_keys(Keys.ENTER)


# -----------------2. Abrir URL correcta -------------------------
driver.get("http://localhost:8000/usuarios")


# ---------- 3. CARGAR EXCEL ----------
workbook = load_workbook("usuarios.xlsx")
sheet = workbook.active

# ---------- 4. RECORRER EXCEL ----------
for fila in range(2, sheet.max_row + 1):

    username = sheet.cell(row=fila, column=1).value
    email = sheet.cell(row=fila, column=2).value
    password = sheet.cell(row=fila, column=3).value


    if not username or not email or not password:
        continue

    print(f"🔍 Agregando: {username}")

    
    driver.get("http://localhost:8000/agregar_usuario")

    btn = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, "btn_agregar"))
)
    btn.click()


    # ---------- 5. BUSCAR ----------
    

   

    try:

        input_username = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.NAME, "username"))
        )

        input_email = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.NAME, "email"))
        )

        input_password = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.NAME, "password"))
        )

        input_username.clear()
        input_username.send_keys(username)

        input_email.clear()
        input_email.send_keys(email)

        input_password.clear()
        input_password.send_keys(password)

        input_password.send_keys(Keys.ENTER)


        mensaje = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "alert"))
        )

        texto_mensaje = mensaje.text
        print("📢 Mensaje:", texto_mensaje)



        if "Usuario agregado exitosamente" in driver.page_source:
            print(f"✅ Usuario {username} agregado exitosamente.")
            reporte.write(f"✅ Usuario {username} agregado exitosamente.\n")
            reporte.write("Mensaje de confirmación:\n" + texto_mensaje + "\n")
            reporte.write("------------------------------------------------\n")


        else:
            print(f"❌ Error al agregar usuario {username}.")
            reporte.write(f"❌ Error al agregar usuario {username}.\n")
            reporte.write("Mensaje de error:\n" + texto_mensaje + "\n")
            reporte.write("------------------------------------------------\n")

    except Exception as e:
        print(f"❌ Ocurrió un error: {str(e)}")
        reporte.write(f"❌ Ocurrió un error al intentar agregar usuario {username}.\n")
        reporte.write("Error:\n" + str(e) + "\n")
        reporte.write("------------------------------------------------\n")
    
    time.sleep(2)


reporte.write("=== FIN DEL REPORTE ===\n")
reporte.close()
driver.quit() 
    
print("✅ Prueba de Agregar Usuarios terminada")
