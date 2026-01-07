from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook
import time

# ---------- 1. ABRIR NAVEGADOR ----------
driver = webdriver.Chrome()
driver.get("http://localhost:8000/iniciarSesion")

username = driver.find_element(By.NAME, "username")
password = driver.find_element(By.NAME, "password")

username.send_keys("je2006")
password.send_keys("Manila2026")
password.send_keys(Keys.ENTER)



driver.get("http://localhost:8000/usuarios")

# ---------- 2. CARGAR EXCEL ----------
workbook = load_workbook("usuarios.xlsx")
sheet = workbook.active

# ---------- 3. RECORRER EXCEL ----------
for fila in range(2, sheet.max_row + 1):

    valor_busqueda = sheet.cell(row=fila, column=1).value

    if not valor_busqueda:
        break

    print(f"🔍 Buscando: {valor_busqueda}")

    # ---------- 4. BUSCAR ----------
    busqueda = driver.find_element(By.NAME, "buscador_Users")
    if not busqueda:
        print("❌ Campo de búsqueda no encontrado")
        continue
    busqueda.clear()
 
    busqueda.send_keys(valor_busqueda)
    busqueda.send_keys(Keys.ENTER)

# 1️⃣ Buscar fila y botón
    fila = driver.find_element(
    By.XPATH,
    f"//tr[td[contains(text(), '{valor_busqueda}')]]"
    )
    boton = fila.find_element(By.TAG_NAME, "button")

    estado_antes = boton.text
    boton.click()

# 2️⃣ Esperar que el DOM se actualice
    WebDriverWait(driver, 10).until(
        EC.staleness_of(boton)
    )

# 3️⃣ VOLVER A BUSCAR la fila
    fila = driver.find_element(
        By.XPATH,
        f"//tr[td[contains(text(), '{valor_busqueda}')]]"
    )
    boton = fila.find_element(By.TAG_NAME, "button")

    estado_despues = boton.text
    print("Estado actual:", estado_despues)


    # ---------- 5. ESPERAR RESULTADO ----------
    WebDriverWait(driver, 10).until(
        EC.text_to_be_present_in_element(
            (By.XPATH, "//tr[td[contains(text(), '{valor}')]]//button".format(valor=valor_busqueda)),
        "Activar"
        )
    )

    time.sleep(2)  # Esperar a que los resultados se carguen
    

print("✅ Prueba de búsqueda terminada")
