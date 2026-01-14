from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook
import time



# ---------- INICIO REPORTE .txt ----------

reporte = open("agregar_etb.txt", "w", encoding="utf-8")
reporte.write("=== REPORTE AUTOMATIZACIÓN SELENIUM ===\n\n")

###################### Iniciacion del programa e inicio de sesion ############################

# ---------- 1. ABRIR NAVEGADOR ----------
driver = webdriver.Chrome()
driver.get("http://localhost:8000/iniciarSesion")

username = driver.find_element(By.NAME, "username")
password = driver.find_element(By.NAME, "password")

username.send_keys("emergia") #Aca va el usuario
password.send_keys("Manila2026") #Aca va la contraseña
password.send_keys(Keys.ENTER) 

# ---------- 2. IR A ETB usuarios ----------    

# -----------------2. Abrir URL correcta -------------------------
driver.get("http://localhost:8000/usuarios_etb")

btn = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, "btn_agregar"))
)
btn.click()

# ---------- 3. CARGAR EXCEL ----------

# ---------- 3. CARGAR EXCEL ----------
workbook = load_workbook("usuariosetb.xlsx")
sheet = workbook.active




for fila in range(2, sheet.max_row + 1):

    clasificacion = sheet.cell(row=fila, column=1).value
    login = sheet.cell(row=fila, column=2).value
    estado = sheet.cell(row=fila, column=3).value

    identificacion_usuario = sheet.cell(row=fila, column=4).value
    fecha_ingreso = sheet.cell(row=fila, column=5).value
    fecha_retiro = sheet.cell(row=fila, column=6).value

    vicepresidente = sheet.cell(row=fila, column=7).value
    gerencia = sheet.cell(row=fila, column=8).value
    direccion = sheet.cell(row=fila, column=9).value

    canal = sheet.cell(row=fila, column=10).value
    proovedor = sheet.cell(row=fila, column=11).value
    nro_contrato = sheet.cell(row=fila, column=12).value

    subcanal = sheet.cell(row=fila, column=13).value
    punto = sheet.cell(row=fila, column=14).value
    cc_supervisor = sheet.cell(row=fila, column=15).value

    lider_etb = sheet.cell(row=fila, column=16).value
    departamento = sheet.cell(row=fila, column=17).value
    ciudad = sheet.cell(row=fila, column=18).value

    nombres = sheet.cell(row=fila, column=19).value
    apellidos = sheet.cell(row=fila, column=20).value
    correo = sheet.cell(row=fila, column=21).value

    celular = sheet.cell(row=fila, column=22).value
    eps = sheet.cell(row=fila, column=23).value
    fecha_nacimiento = sheet.cell(row=fila, column=24).value

    sexo = sheet.cell(row=fila, column=25).value
    cargo = sheet.cell(row=fila, column=26).value
    perfil_portal_suma = sheet.cell(row=fila, column=27).value

    restriccion = sheet.cell(row=fila, column=28).value
    causa_restriccion = sheet.cell(row=fila, column=29).value
    id_fibra = sheet.cell(row=fila, column=30).value

    fija = sheet.cell(row=fila, column=31).value
    movil = sheet.cell(row=fila, column=32).value
    id_cobre = sheet.cell(row=fila, column=33).value

    if not clasificacion or not login or not estado or not identificacion_usuario or not fecha_ingreso or not fecha_retiro or not vicepresidente  or not gerencia  or not direccion or not canal or not proovedor or not nro_contrato or not subcanal or not punto or not cc_supervisor or not lider_etb or not departamento or not ciudad or not nombres or not apellidos or not correo or not celular or not eps or not fecha_nacimiento or not sexo or not cargo or not perfil_portal_suma or not restriccion or not causa_restriccion or not id_fibra or not fija or not movil or not id_cobre:
        continue


    print(f"🔍 Agregando: {username}")

    
    driver.get("http://localhost:8000/agregar_usuario")

    btn = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CLASS_NAME, "btn_agregar"))
)
    btn.click()


    # ---------- 5. BUSCAR ----------
    

###### basarse en el name que esta definido
###### tambien tener encuenta el  / EC.visibility_of_element_located

    try:
        
        input_clasificacion = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "clasificacion"))
        )


        input_login = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "login"))
        )

        input_estado = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "estado"))
        )
########################################
        input_identificacion_usuario = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "identificacion_usuario"))
        )


        input_fecha_ingreso = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "fecha_ingreso"))
        )

        input_fecha_retiro = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "fecha_retiro"))
        )
#####################################
        input_vicepresidente = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "vicepresidente"))
        )


        input_gerencia = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "gerencia"))
        )

        input_direccion = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "direccion"))
        )
######################################
        input_canal = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "canal"))
        )


        input_proveedor = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "proveedor"))
        )

        input_nro_contrato = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "nro_contrato"))
        )

######################################
        input_subcanal = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "subcanal"))
        )

        input_punto = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "punto"))
        )

        input_cc_supervisor = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "cc_supervisor"))
        )
######################################

        input_lider_etb = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "lider_etb"))
        )

        input_departamento = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "departamento"))
        )

        input_ciudad = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "ciudad"))
        )

######################################

        input_nombres = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "nombres"))
        )

        input_apellidos = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "apellidos"))
        )

        input_correo = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "correo"))
        )


######################################

        input_celular = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "celular"))
        )

        input_eps = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "eps"))
        )

        input_fecha_nacimiento = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "fecha_nacimiento"))
        )

######################################

        input_sexo = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "sexo"))
        )

        input_cargo = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "cargo"))
        )

        input_perfil_portal_suma = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "perfil_portal_suma"))
        )
######################################

        input_restriccion = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "restriccion"))
        )

        input_causa_restriccion = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "causa_restriccion"))
        )

        input_id_fibra = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "id_fibra"))
        )


######################################

        input_fija = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "fija"))
        )

        input_movil = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "movil"))
        )

        input_id_cobre = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.NAME, "id_cobre"))
        )



#######################################

        input_clasificacion.clear()
        input_clasificacion.send_keys(clasificacion)

        input_login.clear()
        input_login.send_keys(login)

        input_estado.clear()
        input_estado.send_keys(estado)

######################################

        input_identificacion_usuario.clear()
        input_identificacion_usuario.send_keys(identificacion_usuario)

        input_fecha_ingreso.clear()
        input_login.send_keys(fecha_ingreso)

        input_fecha_retiro.clear()
        input_fecha_retiro.send_keys(fecha_retiro)

#######################################

        input_vicepresidente.clear()
        input_vicepresidente.send_keys(vicepresidente)

        input_gerencia.clear()
        input_gerencia.send_keys(gerencia)

        input_direccion.clear()
        input_direccion.send_keys(direccion)
#######################################
        input_canal.clear()
        input_canal.send_keys(canal)

        input_proveedor.clear()
        input_proveedor.send_keys(proveedor)

        input_nro_contrato.clear()
        input_nro_contrato.send_keys(nro_contrato)
#######################################
        input_subcanal.clear()
        input_subcanal.send_keys(subcanal)

        input_punto.clear()
        input_punto.send_keys(punto)

        input_cc_supervisor.clear()
        input_cc_supervisor.send_keys(cc_supervisor)

#######################################
        input_lider_etb.clear()
        input_lider_etb.send_keys(lider_etb)

        input_departamento.clear()
        input_departamento.send_keys(departamento)

        input_ciudad.clear()
        input_ciudad.send_keys(ciudad)
#######################################
        input_nombres.clear()
        input_nombres.send_keys(nombres)

        input_apellidos.clear()
        input_apellidos.send_keys(apellidos)

        input_correo.clear()
        input_correo.send_keys(correo)
#######################################
        input_celular.clear()
        input_celular.send_keys(celular)

        input_eps.clear()
        input_eps.send_keys(eps)

        input_fecha_nacimiento.clear()
        input_fecha_nacimiento.send_keys(fecha_nacimiento)
#######################################
        input_sexo.clear()
        input_sexo.send_keys(sexo)

        input_cargo.clear()
        input_cargo.send_keys(cargo)

        input_perfil_portal_suma.clear()
        input_perfil_portal_suma.send_keys(perfil_portal_suma)

#######################################
        input_restriccion.clear()
        input_restriccion.send_keys(restriccion)

        input_causa_restriccion.clear()
        input_causa_restriccion.send_keys(causa_restriccion)

        input_id_fibra.clear()
        input_id_fibra.send_keys(id_fibra)

#######################################
        input_fija.clear()
        input_fija.send_keys(fija)

        input_movil.clear()
        input_movil.send_keys(movil)

        input_id_cobre.clear()
        input_id_cobre.send_keys(id_cobre)


################## este input dejarlo par alo ultimo ######################

        input_id_cobre.send_keys(Keys.ENTER)


        mensaje = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "alert"))
        )

        texto_mensaje = mensaje.text
        print("📢 Mensaje:", texto_mensaje)



        if "Usuario agregado exitosamente" in driver.page_source:
            print(f"✅ Usuario {nombres} agregado exitosamente.")
            reporte.write(f"✅ Usuario {nombres} agregado exitosamente.\n")
            reporte.write("Mensaje de confirmación:\n" + texto_mensaje + "\n")
            reporte.write("------------------------------------------------\n")


        else:
            print(f"❌ Error al agregar usuario {nombres}.")
            reporte.write(f"❌ Error al agregar usuario {nombres}.\n")
            reporte.write("Mensaje de error:\n" + texto_mensaje + "\n")
            reporte.write("------------------------------------------------\n")

    except Exception as e:
        print(f"❌ Ocurrió un error: {str(e)}")
        reporte.write(f"❌ Ocurrió un error al intentar agregar usuario {nombres}.\n")
        reporte.write("Error:\n" + str(e) + "\n")
        reporte.write("------------------------------------------------\n")
    
    time.sleep(2)


reporte.write("=== FIN DEL REPORTE ===\n")
reporte.close()
driver.quit() 
    
print("✅ Prueba de Agregar Usuarios ETB terminada")

