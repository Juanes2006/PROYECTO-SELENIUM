from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook
import time


# ---------------- INICIALIZAR ARCHIVO REPORTE .txt   ------------------------
reporte = open("reporte_busqueda_usuarios.txt", "w", encoding="utf-8")
reporte.write("=== REPORTE BUSQUEDA DE USUARIOS ===\n\n")




# ---------- 1. ABRIR NAVEGADOR ----------
driver = webdriver.Chrome()
driver.get("http://localhost:8000/iniciarSesion")

username = driver.find_element(By.NAME, "username")
password = driver.find_element(By.NAME, "password")

username.send_keys("je2006")
password.send_keys("Manila2026")
password.send_keys(Keys.ENTER)



driver.get("http://localhost:8000/usuarios")

# ---------- 2. CARGAR EXCEL ----------
workbook = load_workbook("usuarios.xlsx")
sheet = workbook.active

# ---------- 3. RECORRER EXCEL ----------
for fila in range(2, sheet.max_row + 1):

    valor_busqueda = sheet.cell(row=fila, column=1).value

    if not valor_busqueda:
        break

    print(f"🔍 Buscando: {valor_busqueda}")

    # ---------- 4. BUSCAR ----------
    busqueda = driver.find_element(By.NAME, "buscador_Users")
    busqueda.clear()
    busqueda.send_keys(valor_busqueda)


    try:
        busqueda.send_keys(Keys.ENTER)

    # ---------- 5. ESPERAR RESULTADO ----------
        WebDriverWait(driver, 10).until(
        EC.url_contains("buscador_Users={valor_busqueda}".format(valor_busqueda=valor_busqueda)))
        
        mensaje = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "alert"))
        )

        texto_mensaje = mensaje.text
        print("📢 Mensaje:", texto_mensaje)

        if "Se encontraron 1 usuarios que coinciden con la búsqueda." in driver.page_source:
            print(f"✅ Usuario {valor_busqueda} encontrado exitosamente.")
            reporte.write(f"✅ Usuario {valor_busqueda} encontrado exitosamente.\n")
            reporte.write("Mensaje de confirmación:\n" + texto_mensaje + "\n")
            reporte.write("------------------------------------------------\n")
            
           
        else:
            print(f"❌ Error al encontrar usuario {valor_busqueda}.")
            reporte.write(f"❌ Error al encontrar usuario {valor_busqueda}.\n")
            reporte.write("Mensaje de error:\n" + texto_mensaje + "\n")
            reporte.write("------------------------------------------------\n")
            
    
    except Exception as e:
        print(f"❌ Ocurrió un error: {str(e)}")
        reporte.write(f"❌ Ocurrió un error al intentar encontrar usuario {valor_busqueda}.\n")
        reporte.write("Error:\n" + str(e) + "\n")
        reporte.write("------------------------------------------------\n")

        
    time.sleep(4)  # Esperar a que los resultados se carguen

reporte.write("\n=== FIN REPORTE BUSQUEDA DE USUARIOS ===\n")
reporte.close()
driver.quit()

print("✅ Prueba de búsqueda terminada")
