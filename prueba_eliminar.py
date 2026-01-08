from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook
import time


# ---------- INICIO REPORTE .txt ----------

reporte = open("reporte_eliminar_usuarios.txt", "w", encoding="utf-8")
reporte.write("=== REPORTE ELIMINACION DE USUARIOS ===\n\n")



# ---------- 1. ABRIR NAVEGADOR ----------
driver = webdriver.Chrome()
driver.get("http://localhost:8000/iniciarSesion")

username = driver.find_element(By.NAME, "username")
password = driver.find_element(By.NAME, "password")

username.send_keys("je2006")
password.send_keys("Manila2026")
password.send_keys(Keys.ENTER)



# --------------- 2. Abrir url correcta ------------------------------- 

driver.get("http://localhost:8000/usuarios")


# ---------- 3. CARGAR EXCEL ----------
workbook = load_workbook("usuarios.xlsx")
sheet = workbook.active



# ---------- 4. RECORRER EXCEL ----------
for row in range(2, sheet.max_row + 1):

    valor_busqueda = sheet.cell(row=row, column=1).value

    if not valor_busqueda:
        break

    print(f"🔍 Buscando: {valor_busqueda}")

    busqueda = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.NAME, "buscador_Users"))
    )

    busqueda.clear()
    busqueda.send_keys(str(valor_busqueda))
    busqueda.send_keys(Keys.ENTER)

    WebDriverWait(driver, 10).until(
        EC.url_contains(str(valor_busqueda))
    )

    filas = driver.find_elements(
        By.XPATH,
        f"//tr[td[contains(text(), '{valor_busqueda}')]]"
    )

    if not filas:
        print("❌ Usuario no encontrado")
        continue

    fila_usuario = filas[0]
    boton_eliminar = fila_usuario.find_element(
    By.XPATH, ".//button[contains(text(), 'Eliminar')]"
)
    try:
        boton_eliminar.click()

        mensaje = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "alert"))
        )

        texto_mensaje = mensaje.text
        print("📢 Mensaje:", texto_mensaje)

        if "Usuario eliminado correctamente." in driver.page_source:
            print(f"✅ Usuario {valor_busqueda} eliminado exitosamente.")
            reporte.write(f"✅ Usuario {valor_busqueda} ELIMINADO exitosamente.\n")
            reporte.write("Mensaje de confirmación:\n" + texto_mensaje + "\n")
            reporte.write("------------------------------------------------\n")
            
           
        else:
            print(f"❌ Error al eliminar usuario {valor_busqueda}.")
            reporte.write(f"❌ Error al eliminar usuario {valor_busqueda}.\n")
            reporte.write("Mensaje de error:\n" + texto_mensaje + "\n")
            reporte.write("------------------------------------------------\n")
            

    except Exception as e:
        print(f"❌ Ocurrió un error: {str(e)}")
        reporte.write(f"❌ Ocurrió un error al intentar eliminar usuario {valor_busqueda}.\n")
        reporte.write("Error:\n" + str(e) + "\n")
        reporte.write("------------------------------------------------\n")

    time.sleep(2)


reporte.write("\n=== FIN REPORTE ELIMINACION DE USUARIOS ===\n")
reporte.close()
driver.quit()





print("✅ Prueba de búsqueda terminada")
